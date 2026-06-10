"""
knowledge.py — Base de Conocimiento RAG por empresa.

Endpoints:
  GET    /api/knowledge/           lista paginada de documentos
  POST   /api/knowledge/upload     sube texto/PDF, chunking + embeddings
  DELETE /api/knowledge/{doc_id}   elimina un documento y sus chunks
  GET    /api/knowledge/search     búsqueda semántica de prueba
  GET    /api/knowledge/external   config BD externa
  POST   /api/knowledge/external   guarda config BD externa
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import re
from typing import Any

import aiohttp
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form

from services.supabase_service import supabase, sb_query
from services.auth import CurrentUser, require_admin, get_current_user
from services.embedding_service import get_embedding, search_knowledge, _split_into_chunks
from services.crypto_service import encrypt_data, decrypt_data
from services.document_parser import parse_services_excel, stringify_json_document

logger = logging.getLogger("api-backend")

router = APIRouter(prefix="/api/knowledge", tags=["knowledge"])


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _empresa_id_for_user(user: CurrentUser) -> int:
    """Devuelve empresa_id del usuario; superadmin debe pasarlo como query param."""
    if user.role in ("superadmin",):
        return 0
    return int(user.empresa_id or 0)


def _resolve_empresa(user: CurrentUser, empresa_id_param: int | None) -> int:
    if user.role in ("superadmin",) and empresa_id_param:
        return empresa_id_param
    return int(user.empresa_id or 0)


def _apply_agent_scope(query, agent_id: int | None):
    """NULL agent_id = documentos compartidos de la empresa."""
    if agent_id is not None:
        return query.eq("agent_id", agent_id)
    return query.is_("agent_id", "null")


# ─────────────────────────────────────────────────────────────────────────────
# LISTA
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/company-context")
async def get_company_context(
    empresa_id: int | None = Query(None),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Contexto textual compartido de la empresa (no por agente)."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")
    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")
    res = await sb_query(
        lambda: supabase.table("empresas")
        .select("id, nombre, company_context, kb_allow_internet_search")
        .eq("id", eid)
        .limit(1)
        .execute()
    )
    row = (res.data or [None])[0]
    if not row:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return {
        "empresa_id": eid,
        "nombre": row.get("nombre"),
        "company_context": row.get("company_context") or "",
        "kb_allow_internet_search": bool(row.get("kb_allow_internet_search")),
    }


@router.put("/company-context")
async def save_company_context(
    payload: dict,
    current_user: CurrentUser = Depends(require_admin),
):
    """Guarda el contexto de empresa compartido por todos los agentes."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")
    eid = _resolve_empresa(current_user, payload.get("empresa_id"))
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")
    context = str(payload.get("company_context") or "").strip()
    update_payload: dict[str, Any] = {"company_context": context}
    if "kb_allow_internet_search" in payload:
        update_payload["kb_allow_internet_search"] = bool(payload.get("kb_allow_internet_search"))
    await sb_query(
        lambda: supabase.table("empresas")
        .update(update_payload)
        .eq("id", eid)
        .execute()
    )
    return {
        "status": "ok",
        "empresa_id": eid,
        "company_context": context,
        "kb_allow_internet_search": update_payload.get("kb_allow_internet_search"),
    }


@router.get("/")
async def list_knowledge(
    empresa_id: int | None = Query(None),
    agent_id: int | None = Query(None, description="Si se omite, solo docs de empresa (agent_id NULL)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Lista documentos de la base de conocimiento (paginado)."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    offset = (page - 1) * page_size

    # Agrupamos por (titulo, source_type) para mostrar un doc = todos sus chunks
    res = await sb_query(
        lambda eid=eid, aid=agent_id, lim=page_size, off=offset: _apply_agent_scope(
            supabase.table("knowledge_base")
            .select("id, titulo, source_type, chunk_index, created_at, updated_at")
            .eq("empresa_id", eid),
            aid,
        )
        .order("created_at", desc=True)
        .range(off, off + lim - 1)
        .execute()
    )

    # Agrupar por título para contar chunks
    docs: dict[str, dict] = {}
    for row in res.data or []:
        key = f"{row['titulo']}|{row['source_type']}"
        if key not in docs:
            docs[key] = {
                "titulo": row["titulo"],
                "source_type": row["source_type"],
                "chunks": 0,
                "created_at": row["created_at"],
                "ids": [],
            }
        docs[key]["chunks"] += 1
        docs[key]["ids"].append(row["id"])

    return list(docs.values())


# ─────────────────────────────────────────────────────────────────────────────
# UPLOAD
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/upload", status_code=201)
async def upload_knowledge(
    file: UploadFile = File(...),
    titulo: str = Form(...),
    source_type: str = Form("manual"),
    empresa_id: int | None = Form(None),
    agent_id: int | None = Form(None),
    current_user: CurrentUser = Depends(require_admin),
):
    """
    Sube un documento (PDF o texto), lo divide en chunks de 800 tokens
    con 100 de solapamiento, genera embeddings y los inserta en knowledge_base.
    """
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    # Leer contenido del archivo
    content_bytes = await file.read()
    filename = file.filename or ""
    content_type = (file.content_type or "").lower()

    parsed_documents: list[dict[str, Any]] = []
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext in {"xlsx", "xls"}:
        parsed_documents = await parse_services_excel(content_bytes, eid)

    texto = _extract_text_from_uploaded_file(filename, content_type, content_bytes)
    if parsed_documents:
        texto = "\n\n".join(
            str(doc.get("contenido") or "").strip()
            for doc in parsed_documents
            if str(doc.get("contenido") or "").strip()
        )

    texto = texto.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="El archivo no contiene texto extraíble")

    # Chunkear
    chunks = _split_into_chunks(texto, max_tokens=800, overlap=100)
    if not chunks:
        raise HTTPException(status_code=400, detail="No se pudieron generar chunks del texto")

    # Generar embeddings y preparar filas (en paralelo, lotes de 5)
    rows_to_insert: list[dict] = []
    BATCH = 5
    for i in range(0, len(chunks), BATCH):
        batch = chunks[i:i + BATCH]
        embeddings = await asyncio.gather(
            *[get_embedding(c) for c in batch], return_exceptions=True
        )
        for j, (chunk, emb) in enumerate(zip(batch, embeddings)):
            if isinstance(emb, Exception) or emb is None:
                logger.warning("[knowledge] Embedding fallido chunk %d, se inserta sin vector", i + j)
                emb = None
            row: dict = {
                "empresa_id": eid,
                "titulo": titulo,
                "contenido": chunk,
                "chunk_index": i + j,
                "embedding": emb,
                "source_type": source_type,
            }
            if agent_id is not None:
                row["agent_id"] = int(agent_id)
            rows_to_insert.append(row)

    # Insertar en Supabase
    try:
        res = await sb_query(
            lambda rows=rows_to_insert: supabase.table("knowledge_base").insert(rows).execute()
        )
    except Exception as ins_err:
        logger.error("[knowledge] Error insertando chunks: %s", ins_err)
        raise HTTPException(status_code=500, detail=f"Error al guardar en la base de datos: {ins_err}")

    inserted = len(res.data) if res.data else 0
    logger.info(
        "[knowledge] Documento '%s' indexado: %d chunks, %d con embedding para empresa %d",
        titulo, len(chunks), sum(1 for r in rows_to_insert if r["embedding"]), eid,
    )
    return {
        "status": "ok",
        "titulo": titulo,
        "chunks_total": len(chunks),
        "chunks_con_embedding": sum(1 for r in rows_to_insert if r["embedding"]),
        "insertados": inserted,
        "preview": [
            {
                "titulo": doc.get("titulo"),
                "contenido_preview": str(doc.get("contenido") or "")[:240],
            }
            for doc in parsed_documents[:10]
        ],
    }


@router.post("/upload-url", status_code=201)
async def upload_knowledge_url(
    url: str = Form(...),
    titulo: str = Form(...),
    source_type: str = Form("web"),
    empresa_id: int | None = Form(None),
    agent_id: int | None = Form(None),
    current_user: CurrentUser = Depends(require_admin),
):
    """Indexa una URL remota como documento de la base de conocimiento."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    safe_url = (url or "").strip()
    if not re.match(r"^https?://", safe_url, flags=re.IGNORECASE):
        raise HTTPException(status_code=400, detail="La URL debe empezar por http:// o https://")

    texto = await _extract_text_from_url(safe_url)
    texto = texto.strip()
    if not texto:
        raise HTTPException(status_code=400, detail="No se pudo extraer contenido de la URL")

    chunks = _split_into_chunks(texto, max_tokens=800, overlap=100)
    if not chunks:
        raise HTTPException(status_code=400, detail="No se pudieron generar chunks del texto")

    rows_to_insert: list[dict] = []
    BATCH = 5
    for i in range(0, len(chunks), BATCH):
        batch = chunks[i:i + BATCH]
        embeddings = await asyncio.gather(
            *[get_embedding(c) for c in batch], return_exceptions=True
        )
        for j, (chunk, emb) in enumerate(zip(batch, embeddings)):
            if isinstance(emb, Exception) or emb is None:
                emb = None
            row: dict = {
                "empresa_id": eid,
                "titulo": titulo,
                "contenido": chunk,
                "chunk_index": i + j,
                "embedding": emb,
                "source_type": source_type or "web",
            }
            if agent_id is not None:
                row["agent_id"] = int(agent_id)
            rows_to_insert.append(row)

    res = await sb_query(
        lambda rows=rows_to_insert: supabase.table("knowledge_base").insert(rows).execute()
    )
    return {
        "status": "ok",
        "titulo": titulo,
        "chunks_total": len(chunks),
        "chunks_con_embedding": sum(1 for r in rows_to_insert if r["embedding"]),
        "insertados": len(res.data) if res.data else 0,
    }


def _extract_text_from_pdf(content: bytes) -> str:
    """Extrae texto de un PDF. Intenta pypdf, luego pymupdf como fallback."""
    try:
        import pypdf  # type: ignore
        import io
        reader = pypdf.PdfReader(io.BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
        return "\n".join(pages)
    except ImportError:
        pass

    try:
        import fitz  # type: ignore (PyMuPDF)
        doc = fitz.open(stream=content, filetype="pdf")
        pages = [page.get_text() for page in doc]
        return "\n".join(pages)
    except ImportError:
        pass

    logger.warning("[knowledge] No hay librería PDF disponible (pypdf / PyMuPDF). Instala una de ellas.")
    return ""


def _extract_text_from_excel(content: bytes) -> str:
    """Extrae texto de Excel (.xlsx/.xls) como líneas tabuladas por fila."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        logger.warning("[knowledge] openpyxl no instalado: no se puede procesar Excel")
        return ""

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            lines.append(f"### Hoja: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_vals = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if row_vals:
                    lines.append(" | ".join(row_vals))
        return "\n".join(lines)
    except Exception as exc:
        logger.warning("[knowledge] Error leyendo Excel: %s", exc)
        return ""


def _extract_text_from_uploaded_file(filename: str, content_type: str, content_bytes: bytes) -> str:
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    is_pdf = "pdf" in content_type or ext == "pdf"
    is_excel = ext in {"xlsx", "xls"} or "spreadsheet" in content_type or "excel" in content_type
    is_csv = ext == "csv" or "csv" in content_type
    is_json = ext == "json" or "json" in content_type
    is_docx = ext == "docx" or "word" in content_type

    if is_pdf:
        return _extract_text_from_pdf(content_bytes)
    if is_excel:
        return _extract_text_from_excel(content_bytes)
    if is_docx:
        try:
            from docx import Document  # type: ignore

            doc = Document(io.BytesIO(content_bytes))
            return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except Exception:
            return ""
    if is_csv:
        try:
            decoded = content_bytes.decode("utf-8", errors="replace")
            rows = list(csv.reader(io.StringIO(decoded)))
            return "\n".join(" | ".join(cell.strip() for cell in row if cell and cell.strip()) for row in rows)
        except Exception:
            return ""
    if is_json:
        try:
            parsed = json.loads(content_bytes.decode("utf-8", errors="replace"))
            return stringify_json_document(parsed)
        except Exception:
            return content_bytes.decode("utf-8", errors="replace")

    try:
        return content_bytes.decode("utf-8", errors="replace")
    except Exception:
        return ""


async def _extract_text_from_url(url: str) -> str:
    """Descarga una URL y extrae texto simple desde HTML."""
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(
                url,
                timeout=aiohttp.ClientTimeout(total=10),
                headers={"User-Agent": "AusartaKB/1.0"},
            ) as resp:
                if resp.status != 200:
                    return ""
                html = await resp.text(errors="replace")
    except Exception:
        return ""

    html = re.sub(r"<script\b[^<]*(?:(?!<\/script>)<[^<]*)*<\/script>", " ", html, flags=re.IGNORECASE)
    html = re.sub(r"<style\b[^<]*(?:(?!<\/style>)<[^<]*)*<\/style>", " ", html, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", html)
    text = re.sub(r"\s+", " ", text).strip()
    return text


# ─────────────────────────────────────────────────────────────────────────────
# DELETE
# ─────────────────────────────────────────────────────────────────────────────

@router.delete("/{titulo_encoded}", status_code=204)
async def delete_knowledge_by_title(
    titulo_encoded: str,
    empresa_id: int | None = Query(None),
    agent_id: int | None = Query(None),
    current_user: CurrentUser = Depends(require_admin),
):
    """Elimina todos los chunks de un documento por título."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    from urllib.parse import unquote
    titulo = unquote(titulo_encoded)

    await sb_query(
        lambda eid=eid, aid=agent_id, t=titulo: _apply_agent_scope(
            supabase.table("knowledge_base")
            .delete()
            .eq("empresa_id", eid)
            .eq("titulo", t),
            aid,
        ).execute()
    )
    return


# ─────────────────────────────────────────────────────────────────────────────
# BÚSQUEDA DE PRUEBA
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/search")
async def test_search(
    q: str = Query(..., min_length=2),
    empresa_id: int | None = Query(None),
    agent_id: int | None = Query(None),
    limit: int = Query(5, ge=1, le=20),
    threshold: float = Query(0.7, ge=0.0, le=1.0),
    current_user: CurrentUser = Depends(get_current_user),
):
    """Prueba la búsqueda semántica en la base de conocimiento."""
    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    results = await search_knowledge(
        eid, q, limit=limit, threshold=threshold, agent_id=agent_id
    )
    return {"query": q, "results": results, "total": len(results)}


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN BD EXTERNA
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/external")
async def get_external_db_config(
    empresa_id: int | None = Query(None),
    current_user: CurrentUser = Depends(require_admin),
):
    """Obtiene la configuración de BD externa de la empresa (sin exponer secrets)."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    res = await sb_query(
        lambda eid=eid: supabase.table("empresa_external_db")
        .select("id, db_type, api_url, api_key_header, queries, activo, created_at, updated_at")
        .eq("empresa_id", eid)
        .limit(1)
        .execute()
    )
    if not res.data:
        return {}
    return res.data[0]


@router.post("/external", status_code=201)
async def save_external_db_config(
    payload: dict,
    empresa_id: int | None = Query(None),
    current_user: CurrentUser = Depends(require_admin),
):
    """Guarda o actualiza la configuración de BD externa. Cifra connection_url y api_key."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Sin conexión a la base de datos")

    eid = _resolve_empresa(current_user, empresa_id)
    if not eid:
        raise HTTPException(status_code=400, detail="empresa_id requerido")

    data: dict[str, Any] = {
        "empresa_id": eid,
        "db_type": payload.get("db_type", "rest"),
        "api_url": payload.get("api_url") or None,
        "api_key_header": payload.get("api_key_header") or "Authorization",
        "queries": payload.get("queries") or {},
        "activo": payload.get("activo", True),
    }

    # Cifrar connection_url si viene
    conn_url = (payload.get("connection_url") or "").strip()
    if conn_url:
        try:
            data["connection_url"] = encrypt_data(conn_url)
        except Exception:
            data["connection_url"] = conn_url

    # Cifrar api_key si viene
    api_key = (payload.get("api_key") or "").strip()
    if api_key:
        try:
            data["api_key_enc"] = encrypt_data(api_key)
        except Exception:
            data["api_key_enc"] = api_key

    # Upsert
    res = await sb_query(
        lambda d=data: supabase.table("empresa_external_db")
        .upsert(d, on_conflict="empresa_id")
        .execute()
    )
    return {"status": "ok", "config": res.data[0] if res.data else {}}
