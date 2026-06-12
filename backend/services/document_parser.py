from __future__ import annotations

import io
import json
from typing import Any

from services.chunk_builder import service_row_to_chunk


async def parse_services_excel(
    file_bytes: bytes,
    empresa_id: int,
    sheet_name: str = "informe CRM",
    agent_id: int | None = None,
    fuente: str = "services_excel",
) -> list[dict[str, Any]]:
    """Convierte cada fila del Excel en un chunk semántico (1 producto = 1 chunk)."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return []

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h or "").strip() for h in rows[0]]
    docs: list[dict[str, Any]] = []
    active_rows = max(0, len(rows) - 1)

    for values in rows[1:]:
        row = {headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))}

        activo = row.get("Activo")
        if activo is not None and activo != "":
            try:
                if float(activo) == 0:
                    continue
            except (TypeError, ValueError):
                if str(activo).strip().lower() in {"0", "no", "false"}:
                    continue

        ofertable = str(row.get("Ofertable") or "").strip().lower()
        activable = str(row.get("Activable") or "").strip().lower()
        if ofertable not in {"si", "sí", "s"} and activable not in {"si", "sí", "s"}:
            continue

        uds_activas = row.get("Uds totales activas")
        try:
            uds_activas_num = float(uds_activas or 0)
        except (TypeError, ValueError):
            uds_activas_num = 0
        if active_rows > 50 and uds_activas_num == 0:
            continue

        chunk = service_row_to_chunk(
            row,
            empresa_id=empresa_id,
            agent_id=agent_id,
            fuente=fuente,
        )
        if chunk:
            docs.append(chunk)

    return docs


async def extract_text_from_docx(file_bytes: bytes) -> str:
    try:
        from docx import Document  # type: ignore
    except ImportError:
        return ""

    doc = Document(io.BytesIO(file_bytes))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


async def normalize_documents_for_preview(documents: list[dict[str, Any]]) -> list[dict[str, Any]]:
    preview: list[dict[str, Any]] = []
    for doc in documents:
        preview.append(
            {
                "titulo": doc.get("titulo") or "",
                "source_type": doc.get("source_type") or "manual",
                "contenido_preview": str(doc.get("contenido") or "")[:500],
            }
        )
    return preview


def stringify_json_document(payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)
