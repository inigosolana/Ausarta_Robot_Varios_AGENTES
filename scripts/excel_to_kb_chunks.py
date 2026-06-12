#!/usr/bin/env python3
"""
Convierte informe_servicios_*.xlsx → chunks.jsonl listos para embeddings.

Uso:
  python scripts/excel_to_kb_chunks.py \\
    --file informe_servicios_20260313.xlsx \\
    --empresa_id 1 \\
    --agente_id 42 \\
    --output data/ausarta_kb_chunks.jsonl

Luego ingesta:
  python scripts/ingest_kb_chunks.py --file data/ausarta_kb_chunks.jsonl
"""
from __future__ import annotations

import argparse
import io
import sys
from pathlib import Path

# Raíz del repo en sys.path para importar backend/
ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend" if (ROOT / "backend" / "config.py").exists() else ROOT
sys.path.insert(0, str(BACKEND))

from services.chunk_builder import service_row_to_chunk, write_jsonl  # noqa: E402


def _read_excel_rows(file_path: str, sheet_name: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl no instalado. Ejecuta: pip install openpyxl") from exc

    wb = load_workbook(file_path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h or "").strip() for h in rows[0]]
    parsed: list[dict] = []
    for values in rows[1:]:
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
        parsed.append(row)
    return parsed


def _row_is_active(row: dict) -> bool:
    activo = row.get("Activo")
    if activo is None or activo == "":
        return True
    try:
        return float(activo) != 0
    except (TypeError, ValueError):
        return str(activo).strip().lower() not in {"0", "no", "false"}


def convert(
    file_path: str,
    empresa_id: int,
    agente_id: int | None,
    output: str,
    sheet_name: str = "informe CRM",
) -> int:
    rows = _read_excel_rows(file_path, sheet_name)
    fuente = Path(file_path).name
    chunks = []

    for row in rows:
        if not _row_is_active(row):
            continue
        ofertable = str(row.get("Ofertable") or "").strip().lower()
        activable = str(row.get("Activable") or "").strip().lower()
        if ofertable not in {"sí", "si", "s"} and activable not in {"sí", "si", "s"}:
            continue

        chunk = service_row_to_chunk(
            row,
            empresa_id=empresa_id,
            agent_id=agente_id,
            fuente=fuente,
        )
        if chunk:
            chunks.append(chunk)

    count = write_jsonl(chunks, output)
    print(f"✅ {count} chunks escritos en {output}")
    return count


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Excel de servicios → JSONL chunks")
    parser.add_argument("--file", required=True, help="Ruta al .xlsx")
    parser.add_argument("--empresa_id", type=int, required=True)
    parser.add_argument("--agente_id", type=int, default=None)
    parser.add_argument("--output", default="data/ausarta_kb_chunks.jsonl")
    parser.add_argument("--sheet", default="informe CRM")
    args = parser.parse_args()
    convert(args.file, args.empresa_id, args.agente_id, args.output, args.sheet)
